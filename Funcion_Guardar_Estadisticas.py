import openpyxl
import matplotlib.pyplot as grafico
import pandas as pd
from IPython.display import display, HTML
from datetime import datetime

def registrar(modo,resultado): # Guarda la partida en el excel inmediatamente después del anterior registro
    archivo_excel = "C:\\Users\\Obadic\\Tabla_partidas.xlsx"
    try: # En caso de que la ruta al archivo excel no sea la correcta
        excel = openpyxl.load_workbook(archivo_excel)
        hoja_excel = excel["Historial_de_partidas"]
    except FileNotFoundError:
        print(f"No se encontró el archivo: {archivo_excel}. No se puedo guardar las estadísticas de la partida.")
        return
    nombre=input("Nombre del jugador: ").strip() # Elimina espacios, para evitar nombres de "espacio espacio" que pueden dar problemas al luego pedir datos/estadísticas
    while not nombre: # Evalúa si nombre tiene algún valor, si no lo tiene, lo pide (evitar nombres en blanco)
        display(HTML('<span style="color:red; font-weight:bold;"><br>El nombre no puede estar vacío. Por favor, introduzca un nombre válido.</span>'))
        nombre = input("Nombre del jugador: ").strip()
    fecha_partida = datetime.now().strftime("%Y-%m-%d %H:%M:%S") # Guarda fecha y hora de la partida
    datos_partida=[nombre,modo,resultado,fecha_partida]
    hoja_excel.append(datos_partida) 
    excel.save("C:\\Users\\Obadic\\Tabla_partidas.xlsx")
    print("¡Los datos de la partida se guardaron correctamente!\n")
    
def mostrar_resultados(): # Visualización de las estadísticas en gráficos o en texto
    archivo_excel = "C:\\Users\\Obadic\\Tabla_partidas.xlsx"
    try: # Try/except para ruta de archivo excel no correcta
        excel = openpyxl.load_workbook(archivo_excel)
        hoja_excel = excel["Historial_de_partidas"]
    except FileNotFoundError:
        print(f"No se encontró el archivo: {archivo_excel}. No se puede acceder a las estadísticas.")
        return
    estadistica_opcion=0
    seguir_menu_estadisticas=True
    while seguir_menu_estadisticas:
        display(HTML('<span style="font-size:16px; font-weight:bold;"><br>Estadísticas:</span>'))
        print("\n1. Historial de partidas\n2. Partidas jugadas por un jugador\n3. Estadísticas de las partidas jugadas\n4. Estadísticas de un jugador\n5. Volver al menú principal\n")
        try: # Try/except por si no se introduce un integer
            estadistica_opcion=(int(input("Introduzca el número de la opción deseada: ")))
        except ValueError:
            display(HTML('<span style="color:red; font-weight:bold;"><br>Debe introducir un número del 1 al 5, en función de la opción que desee elegir.</span>'))
            continue
            
        if estadistica_opcion==1: # Todo el historial de partidas en texto
            print("\nHistorial de todas las partidas:")
            print("-" * 50)
            for fila in hoja_excel.iter_rows(values_only=True): # Itera sobre las filas de la hoja. values_only=True devolve solamente los valores de las celdas, en lugar de los objetos Cell
                print(" | ".join(str(celda) if celda else "" for celda in fila)) # Formato de presentación de la información contenida en el excel
                print()
        elif estadistica_opcion==2: # Historial de partidas filtrada de un jugador en texto
            nombre_partidas=input("Nombre del jugador: ").strip() 
            while not nombre_partidas: # En caso de introducir un nombre en blanco, bucle hasta introducir uno válido
                display(HTML('<span style="color:red; font-weight:bold;"><br>El nombre no puede estar vacío. Por favor, introduzca un nombre válido.</span>'))
                nombre_partidas = input("Nombre del jugador: ").strip()
            print(f"\nResultados de {nombre_partidas}:")
            encontrado=False
            for fila in hoja_excel.iter_rows(values_only=True):
                if fila[0] == nombre_partidas:  # Compara el nombre en la primera columna donde se encuentran los nombres de los jugadores
                    encontrado = True
                    print(f"{fila[0]:<10} | {fila[1]:<12} | {fila[2]:<8} | {fila[3]}")
                    print()
            if not encontrado:
                print(f"No se encontraron partidas para el jugador {nombre_partidas}.")
            
        elif estadistica_opcion == 3: # Gráfico de partidas ganadas/perdidas para todo el historial de partidas
            Ganadas=0
            Perdidas=0
            for fila in hoja_excel.iter_rows(values_only=True): 
                if fila[2] == "Ganada":  
                    Ganadas += 1
                elif fila[2] == "Perdida":  
                    Perdidas += 1
            # Gráfico
            x=["Ganadas", "Perdidas"] # Define X
            y=[Ganadas, Perdidas] # Define Y
            grafico.bar(x,y,color=['green', 'red']) # Color de las barras del gráfico
            grafico.title("Estadísticas de Partidas") # Título
            grafico.xlabel("Resultados") # Etiqueta eje X
            grafico.ylabel("Cantidad") # Etiqueta eje Y
            grafico.show()
            
        elif estadistica_opcion == 4: # Gráfico de partidas ganadas/perdidas filtradas por jugador
            nombre_estadistica=input("Nombre del jugador: ").strip()
            while not nombre_estadistica:
                display(HTML('<span style="color:red; font-weight:bold;"><br>El nombre no puede estar vacío. Por favor, introduzca un nombre válido.</span>'))
                nombre_estadistica = input("Nombre del jugador: ").strip()
            hoja_excel=excel["Historial_de_partidas"]
            datos = pd.read_excel(archivo_excel, sheet_name="Historial_de_partidas")
            datos_jugador = datos[datos["JUGADOR"].str.contains(nombre_estadistica, case=False, na=False)]
                
            if datos_jugador.empty: # En caso de que no se encuentren partidas para el jugador introducido
                print(f"No se encontraron partidas para el jugador: {nombre_estadistica}.")
            else:
                resumen = datos_jugador["RESULTADO"].value_counts()
                    
                # Gráfico
                x = ["Ganadas", "Perdidas"]
                y = [resumen.get("Ganada", 0), resumen.get("Perdida", 0)]
                grafico.bar(x, y, color=["green", "red"])
                grafico.title(f"Resultados de {nombre_estadistica}")
                grafico.xlabel("Resultados")
                grafico.ylabel("Cantidad")
                grafico.show()

        
        elif estadistica_opcion == 5: # Para volver al menú principal
           seguir_menu_estadisticas=False 
        else:
            display(HTML('<span style="color:red; font-weight:bold;"><br>Opción no válida.</span>'))